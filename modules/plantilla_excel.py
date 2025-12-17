from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

class PlantillaExcelProcessor:

    def __init__(self, ruta_plantilla, df_hoja3):
        self.ruta = ruta_plantilla
        self.df_hoja3 = df_hoja3
        self.wb = load_workbook(self.ruta)

    # --------------------------------------------------
    # 1. VOLCADO DE DATOS (Paso Nuevo)
    # --------------------------------------------------
    def escribir_datos_raw(self):
        ws = self.wb["Hoja3"]
        print("→ Volcando datos nuevos en Hoja3...")

        # Limpiar Hoja3 completa (desde fila 2) para evitar residuos
        if ws.max_row > 1:
            ws.delete_rows(2, amount=ws.max_row)

        # Escribir el DataFrame de golpe
        # index=False: No escribe el número de fila de pandas
        # header=False: No repite los títulos, usa los de la plantilla
        for r in dataframe_to_rows(self.df_hoja3, index=False, header=False):
            ws.append(r)
        
        print("✔ Datos volcados exitosamente en Hoja3")

    # --------------------------------------------------
    # 2. TRANSFORMACIONES DE NEGOCIO
    # --------------------------------------------------
    def transformar_hoja_3(self):
        ws = self.wb["Hoja3"]
        print("→ Aplicando reglas de negocio en Hoja3...")

        # Iteramos filas desde la 2.
        # Asumimos estructura: [A=ID, B=Sitio, C=Nombre, D=Estatus, E=Grabación]
        for row in ws.iter_rows(min_row=2):
            if len(row) < 4: continue # Saltar filas vacías
            
            celda_estatus = row[3]   # Columna D
            celda_grabacion = row[4] # Columna E (Donde escribiremos el resultado)

            estatus = str(celda_estatus.value).lower().strip() if celda_estatus.value else ""

            # ---- REGLAS DE ESTATUS ----
            if "con señal" in estatus and "intermitente" not in estatus:
                celda_estatus.value = "Con señal"
                celda_grabacion.value = "Cámara grabando"

            elif "sin señal" in estatus:
                celda_estatus.value = "Sin señal"
                celda_grabacion.value = "Cámara sin grabación"

            elif "intermitente" in estatus:
                celda_estatus.value = "Cámara con señal/intermitente"
                celda_grabacion.value = "Cámara grabando/intermitente"

            elif "desactivada" in estatus:
                celda_estatus.value = "Cámara desactivada"
                celda_grabacion.value = "Cámara sin grabación"

        print("✔ Transformaciones completadas")

    # --------------------------------------------------
    # 3. CONSOLIDACIÓN OPTIMIZADA (TURBO)
    # --------------------------------------------------
    def actualizar_securos(self):
        ws_hoja3 = self.wb["Hoja3"]
        ws_securos = self.wb["securos"]

        print("→ Iniciando limpieza y consolidación INTELIGENTE...")

        # A) Obtener IDs nuevos para no duplicar
        ids_nuevos = set()
        for row in ws_hoja3.iter_rows(min_row=2, values_only=True):
            if row and row[0]: 
                ids_nuevos.add(str(row[0]).strip())

        # B) Leer histórico a MEMORIA (Evita borrar fila por fila que es lento)
        filas_historicas_a_conservar = []
        fila_footer_original = None
        
        # Leemos 'securos' buscando datos viejos y el footer (ID 10321)
        # enumerate empieza en 2 porque es la fila excel real
        for i, row in enumerate(ws_securos.iter_rows(min_row=2, values_only=True), start=2):
            val_id = str(row[0]).strip() if row[0] else ""
            val_cod = row[2] if len(row) > 2 else ""

            # Detectamos el footer para detener la lectura
            if val_id == "10321" or val_cod == "ANPR-CC-10321":
                fila_footer_original = i
                break
            
            # Lógica de Filtrado: Si el ID NO viene en el reporte de hoy, lo guardamos.
            # Si SÍ viene hoy, no lo guardamos (lo reemplazaremos con el nuevo).
            if val_id not in ids_nuevos:
                filas_historicas_a_conservar.append(row)

        print(f"✔ Filtrado en memoria listo. Conservando {len(filas_historicas_a_conservar)} registros históricos.")

        # C) Borrado Masivo del bloque de datos
        # Calculamos cuántas filas borrar antes del footer
        if fila_footer_original:
            cantidad_a_borrar = fila_footer_original - 2
            if cantidad_a_borrar > 0:
                print("→ Borrando bloque de datos antiguos...")
                ws_securos.delete_rows(2, amount=cantidad_a_borrar)
        else:
            # Si no hay footer, borramos todo
            if ws_securos.max_row >= 2:
                ws_securos.delete_rows(2, amount=ws_securos.max_row)

        # D) Reconstrucción: Escribir Histórico Limpio + Datos Nuevos
        print("→ Reescribiendo datos unificados...")
        
        # 1. Escribir Histórico
        fila_actual = 2
        if filas_historicas_a_conservar:
            ws_securos.insert_rows(fila_actual, amount=len(filas_historicas_a_conservar))
            for datos in filas_historicas_a_conservar:
                for col_idx, valor in enumerate(datos, start=1):
                    ws_securos.cell(row=fila_actual, column=col_idx, value=valor)
                fila_actual += 1

        # 2. Escribir Nuevos (Del día)
        filas_nuevas = list(ws_hoja3.iter_rows(min_row=2, values_only=True))
        if filas_nuevas:
            ws_securos.insert_rows(fila_actual, amount=len(filas_nuevas))
            for datos in filas_nuevas:
                for col_idx, valor in enumerate(datos, start=1):
                    ws_securos.cell(row=fila_actual, column=col_idx, value=valor)
                fila_actual += 1

        print("✔ Consolidación FINALIZADA correctamente.")

    # --------------------------------------------------
    # 4. GUARDAR
    # --------------------------------------------------
    def guardar(self):
        try:
            self.wb.save(self.ruta)
            print(f"✔ Archivo guardado ÉXITOSAMENTE: {self.ruta}")
        except PermissionError:
            print("❌ ERROR CRÍTICO: Cierra el archivo Excel antes de ejecutar el script.")